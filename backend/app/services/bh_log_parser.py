"""
General-purpose "report-style" borehole log Excel parser.

Unlike lab_data.py's own template (one flat table, header always row 1) or
universal_soil_parser.py's "flat table somewhere in the first 40 rows"
assumption, a REPORT-style workbook is what most consulting firms actually
print: a fixed-layout page with a title/metadata block, then a data table
whose header can be 1-4 rows tall (main label + unit + sub-label), start
anywhere in the sheet, and span anywhere from ~15 to 250+ columns.

Built 29 Jul 2026 from 6 REAL company report templates Raahi provided
(different labs, different highway/bridge projects) -- not guessed from a
spec. This module replaces an earlier bh_log_parser.py that no longer
existed in the codebase (see PROJECT_STATUS.md entry #37).

HONESTY: this is best-effort pattern matching across genuinely different
templates, not a guarantee. Every result carries which row/column the header
was detected at and a confidence note -- verify against the source PDF/
hardcopy before trusting output for anything safety-relevant, same principle
as universal_soil_parser.py.
"""
from __future__ import annotations
import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app.services.universal_soil_parser import (
    CANONICAL_FIELDS, _normalize, _detect_unit, match_header,
)

MAX_SCAN_ROWS = 80          # how far down to look for the data table's header
MAX_SCAN_COLS = 260         # some of these reports are genuinely this wide
MAX_HEADER_WINDOW = 4        # a header can span up to this many stacked rows
MIN_LAYER_FIELDS_FOR_HEADER = 3   # need at least this many layer-level fields matched to accept a header
MAX_BLANK_ROWS_BEFORE_STOP = 4     # consecutive blank data rows before assuming the table ended

# For borehole ID specifically: text embedded in a title string is common
# ("BOREHOLE NO: BH-01", "BORE HOLE NO. BH-1", "BH No 1") -- regex fallback
# when there's no clean label:value adjacent-cell pair.
_BOREHOLE_ID_PATTERNS = [
    re.compile(r"bore\s*hole\s*no\.?\s*[:\-]?\s*([A-Za-z0-9\-/]+)", re.IGNORECASE),
    re.compile(r"\bbh\s*no\.?\s*[:\-]?\s*([A-Za-z0-9\-/]+)", re.IGNORECASE),
    re.compile(r"\bbh[\-\s]?(\d+[A-Za-z]?)\b", re.IGNORECASE),
]


def _cell_text(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def _is_numeric(v: Any) -> bool:
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v.strip())
            return True
        except ValueError:
            return False
    return False


def _to_float(v: Any) -> Optional[float]:
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def _looks_like_column_index_row(ws, row: int, mapped_cols: List[int]) -> bool:
    """True if the given row's values across the mapped columns are small,
    strictly-increasing integers (1, 2, 3...) -- a decorative "column
    number" row some templates insert right after the real header labels,
    which would otherwise be misread as the first data row. Blank cells
    among the mapped columns are ignored (real index rows often skip
    separator/derived columns); only the populated ones need to fit."""
    vals = []
    for col in sorted(mapped_cols):
        v = ws.cell(row=row, column=col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if not _is_numeric(v):
            return False
        vals.append(_to_float(v))
    if len(vals) < 4:
        return False
    if any(v > 60 or v != int(v) for v in vals):
        return False
    return all(b > a for a, b in zip(vals, vals[1:]))


def _find_data_table(ws, max_row: int, max_col: int) -> Optional[Dict[str, Any]]:
    """
    Two-stage search for the best-scoring header block:
    Stage 1 (cheap): try height=1 at every row up to MAX_SCAN_ROWS, tracking
    which rows match ANY recognized field at all.
    Stage 2 (expensive): only for rows near a Stage-1 hit (start_row-1 to
    start_row+1, since a multi-row header's most-recognizable single row
    isn't always its top row), try header heights 1..MAX_HEADER_WINDOW with
    the full column-by-column concatenated-text matching.
    This two-stage approach is what keeps this fast -- trying every
    (start_row, height) combination blindly across MAX_SCAN_ROWS rows made
    3000+ unique fuzzy-match calls on a real 40-column report sheet (40+
    seconds, causing upload timeouts) before this was added.
    """
    row_scores: Dict[int, int] = {}
    for row in range(1, min(max_row, MAX_SCAN_ROWS) + 1):
        fields_hit = set()
        for col in range(1, min(max_col, MAX_SCAN_COLS) + 1):
            text = _cell_text(ws, row, col)
            if len(text) < 2:
                continue
            match = match_header(text)
            if match and match["confidence"] >= 70 and CANONICAL_FIELDS[match["field"]]["scope"] == "layer":
                fields_hit.add(match["field"])
        if fields_hit:
            row_scores[row] = len(fields_hit)

    if not row_scores:
        search_rows = set(range(1, min(max_row, MAX_SCAN_ROWS) + 1))  # Stage 1 found nothing -- fall back to full scan
    else:
        top_rows = sorted(row_scores, key=row_scores.get, reverse=True)[:5]
        search_rows = set()
        for r in top_rows:
            search_rows.update(range(max(1, r - MAX_HEADER_WINDOW + 1), r + 1))

    best = None
    for start_row in sorted(search_rows):
        for height in range(1, MAX_HEADER_WINDOW + 1):
            if start_row + height - 1 > max_row:
                break
            column_field_map: Dict[int, Dict[str, Any]] = {}
            layer_field_count = 0
            has_depth = False
            for col in range(1, min(max_col, MAX_SCAN_COLS) + 1):
                parts = [_cell_text(ws, r, col) for r in range(start_row, start_row + height)]
                combined = " ".join(p for p in parts if p)
                if not combined or len(combined.strip()) < 2:
                    continue
                match = match_header(combined)
                if not match:
                    continue
                field_key = match["field"]
                info = CANONICAL_FIELDS[field_key]
                if info["scope"] != "layer":
                    continue
                # keep the higher-confidence match if a field key repeats across columns
                existing = column_field_map.get(col)
                if existing and existing["confidence"] >= match["confidence"]:
                    continue
                column_field_map[col] = {"field": field_key, "confidence": match["confidence"], "header_text": combined}
                if field_key in ("from_m", "to_m"):
                    has_depth = True

            if not has_depth:
                # Alternate convention: one shared "Depth ..." header over a
                # From / "-" / To triplet of columns, instead of separate
                # From and To headers (seen in one of the 6 real templates
                # this was built from -- e.g. col=0, col+1='-', col+2=0.5).
                for col in range(1, min(max_col, MAX_SCAN_COLS) + 1):
                    parts = [_cell_text(ws, r, col) for r in range(start_row, start_row + height)]
                    combined = _normalize(" ".join(p for p in parts if p))
                    if "depth" not in combined:
                        continue
                    found_pair = False
                    for probe_row in range(start_row + height, min(start_row + height + 3, max_row) + 1):
                        sep_val = ws.cell(row=probe_row, column=col + 1).value
                        to_val = ws.cell(row=probe_row, column=col + 2).value
                        if isinstance(sep_val, str) and sep_val.strip() == "-" and _is_numeric(to_val):
                            found_pair = True
                            break
                    if found_pair:
                        column_field_map[col] = {"field": "from_m", "confidence": 70.0, "header_text": combined}
                        column_field_map[col + 2] = {"field": "to_m", "confidence": 70.0, "header_text": combined}
                        has_depth = True
                        break

            # de-duplicate: if the SAME field_key won multiple columns, keep only the
            # highest-confidence one (a report table should have one column per field)
            by_field: Dict[str, Tuple[int, float]] = {}
            for col, m in column_field_map.items():
                fk = m["field"]
                if fk not in by_field or m["confidence"] > by_field[fk][1]:
                    by_field[fk] = (col, m["confidence"])
            deduped = {col: column_field_map[col] for col, _ in by_field.values()}
            layer_field_count = len({m["field"] for m in deduped.values()})

            if not has_depth or layer_field_count < MIN_LAYER_FIELDS_FOR_HEADER:
                continue

            data_start = start_row + height
            adjusted_data_start = data_start
            for _ in range(3):
                row_is_blank = all(
                    (lambda v: v is None or (isinstance(v, str) and not v.strip()))(ws.cell(row=adjusted_data_start, column=c).value)
                    for c in deduped.keys()
                )
                if row_is_blank or _looks_like_column_index_row(ws, adjusted_data_start, list(deduped.keys())):
                    adjusted_data_start += 1
                else:
                    break

            depth_cols = [c for c, m in deduped.items() if m["field"] in ("from_m", "to_m")]
            probe_col = depth_cols[0]
            numeric_rows_found = 0
            for r in range(adjusted_data_start, min(adjusted_data_start + 6, max_row + 1)):
                if _is_numeric(ws.cell(row=r, column=probe_col).value):
                    numeric_rows_found += 1
            if numeric_rows_found < 2:
                continue

            score = layer_field_count * 10 + numeric_rows_found
            if best is None or score > best["score"]:
                best = {
                    "score": score, "header_row_start": start_row, "header_height": height,
                    "data_start_row": adjusted_data_start, "column_field_map": deduped,
                }
    return best


def _extract_metadata(ws, max_row: int, max_col: int, stop_row: int) -> Dict[str, Any]:
    """
    Label-scans the area ABOVE the detected data table (where title/metadata
    blocks live in every report format seen) for borehole-level fields --
    project name, borehole ID, water table depth, RL, coordinates, date.
    For each label cell matched, checks a small set of typical value
    positions (same row, a few cells right; or the cell directly below) and
    takes the first one that looks like a plausible value for that field's
    type.
    """
    metadata: Dict[str, Any] = {}
    scan_rows = max(1, stop_row - 1)  # strictly ABOVE the header block, never overlapping it
    for row in range(1, scan_rows + 1):
        for col in range(1, min(max_col, MAX_SCAN_COLS) + 1):
            text = _cell_text(ws, row, col)
            if not text or len(text) > 60:
                continue
            match = match_header(text)
            if not match or match["confidence"] < 90 or CANONICAL_FIELDS[match["field"]]["scope"] != "borehole":
                continue
            field_key = match["field"]
            if field_key == "borehole_id":
                continue  # handled separately below -- label:adjacent-value is unreliable for this one
            if field_key in metadata:
                continue
            candidates = [
                ws.cell(row=row, column=c).value for c in range(col + 1, min(col + 4, max_col) + 1)
            ] + [ws.cell(row=row + 1, column=col).value]
            for cand in candidates:
                if cand is None or (isinstance(cand, str) and not cand.strip()):
                    continue
                if isinstance(cand, str) and match_header(cand):
                    continue  # that's another label, not a value
                metadata[field_key] = cand
                break

    sheet_title = ws.title.strip()
    if re.search(r"\bbh\b|\bbore\s*hole\b|\bhole\b", sheet_title, re.IGNORECASE) and len(sheet_title) <= 20:
        metadata["borehole_id"] = sheet_title

    if "borehole_id" not in metadata:
        for row in range(1, scan_rows + 1):
            for col in range(1, min(max_col, 40) + 1):
                text = _cell_text(ws, row, col)
                if not text:
                    continue
                for pattern in _BOREHOLE_ID_PATTERNS:
                    m = pattern.search(text)
                    if m:
                        metadata["borehole_id"] = m.group(1).strip()
                        break
                if "borehole_id" in metadata:
                    break
            if "borehole_id" in metadata:
                break

    return metadata


def parse_borehole_log_workbook(file_bytes: bytes, sheet_name: Optional[str] = None, _preloaded_wb=None) -> Dict[str, Any]:
    """
    Parses ONE sheet of a report-style workbook (the first sheet, or
    `sheet_name` if given -- multi-sheet workbooks with one borehole per
    sheet need to call this once per sheet, same as universal_soil_parser's
    caller does for "one sheet per borehole" workbooks).

    _preloaded_wb: an already-loaded Workbook, to avoid re-parsing the same
    bytes on every sheet when the caller loops over all sheets (see
    parse_uploaded_workbook_auto -- this used to reload+reparse the entire
    file from scratch once per sheet, which for an N-sheet workbook meant N
    redundant full openpyxl parses of identical bytes; fixed 4 Aug 2026).

    Returns:
    {
        "sheet_name": str,
        "metadata": {borehole-level fields found},
        "layers": [{layer-level fields found, "row": excel_row_number}, ...],
        "header_row_start": int, "header_height": int,
        "warnings": [str, ...],
    }
    Raises ValueError if no data table could be confidently located.
    """
    wb = _preloaded_wb if _preloaded_wb is not None else load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    max_row, max_col = ws.max_row, ws.max_column

    table = _find_data_table(ws, max_row, max_col)
    if table is None:
        raise ValueError(
            f"Sheet '{ws.title}': could not confidently locate a soil-layer data table "
            f"(no header row/column combination matched enough recognized fields with "
            f"plausible depth data beneath it)."
        )

    metadata = _extract_metadata(ws, max_row, max_col, table["header_row_start"])

    layers: List[Dict[str, Any]] = []
    warnings: List[str] = []
    blank_streak = 0
    depth_cols = [c for c, m in table["column_field_map"].items() if m["field"] in ("from_m", "to_m")]
    probe_col = depth_cols[0]

    for row in range(table["data_start_row"], max_row + 1):
        if not _is_numeric(ws.cell(row=row, column=probe_col).value):
            blank_streak += 1
            if blank_streak >= MAX_BLANK_ROWS_BEFORE_STOP:
                break
            continue
        blank_streak = 0

        layer: Dict[str, Any] = {"row": row}
        for col, m in table["column_field_map"].items():
            field_key = m["field"]
            raw_value = ws.cell(row=row, column=col).value
            if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
                continue
            info = CANONICAL_FIELDS[field_key]
            if "unit_target" in info:
                unit = _detect_unit(m["header_text"], field_key)
                numeric = _to_float(raw_value)
                if numeric is not None and unit:
                    numeric *= unit["factor"]
                layer[field_key] = numeric if numeric is not None else raw_value
                if numeric is not None and not unit:
                    warnings.append(
                        f"Row {row}, field '{field_key}': no unit recognized in header "
                        f"'{m['header_text']}' -- value {raw_value} used as-is, verify units."
                    )
            elif field_key in ("from_m", "to_m", "n_value"):
                layer[field_key] = _to_float(raw_value)
            else:
                layer[field_key] = raw_value
        layers.append(layer)

    if not layers:
        raise ValueError(f"Sheet '{ws.title}': header row detected but no data rows found beneath it.")

    # Real report templates sometimes record a SINGLE depth per row (e.g. one
    # SPT test point every 1-1.5m) rather than a From/To range pair. Because
    # "Depth" alone is lexically close to both "depth from" and "depth to"
    # (from_m/to_m's synonyms), match_header picks ONE of them somewhat
    # arbitrarily -- leaving every single layer missing the other required
    # column, which used to mean the whole table came out unusable (entry
    # #42). Detect that (exactly one of from_m/to_m matched in this table,
    # never both) and synthesize the missing boundary from consecutive rows'
    # points instead, flagging it clearly so it gets verified against the
    # source rather than silently trusted.
    matched_fields = {m["field"] for m in table["column_field_map"].values()}
    has_from, has_to = "from_m" in matched_fields, "to_m" in matched_fields
    if has_from != has_to:
        only_field = "from_m" if has_from else "to_m"
        other_field = "to_m" if has_from else "from_m"
        warnings.append(
            f"Only a single depth column was found in this table (matched as '{only_field}', "
            f"header '{[m['header_text'] for m in table['column_field_map'].values() if m['field'] == only_field][0]}') "
            f"-- no separate From/To pair. Treating each row as a POINT depth (e.g. an SPT test "
            f"depth) and synthesizing '{other_field}' from consecutive rows' points. "
            f"VERIFY these synthesized layer boundaries against the source document."
        )
        prev = 0.0
        for layer in layers:
            pt = layer.get(only_field)
            if pt is None:
                continue
            if only_field == "from_m":
                layer["to_m"], layer["from_m"] = pt, prev
            else:
                layer["from_m"], layer["to_m"] = prev, pt
            prev = pt

    return {
        "sheet_name": ws.title,
        "metadata": metadata,
        "layers": layers,
        "header_row_start": table["header_row_start"],
        "header_height": table["header_height"],
        "confidence_note": (
            f"Header detected at row {table['header_row_start']} "
            f"(height {table['header_height']}), {len(table['column_field_map'])} field(s) matched, "
            f"{len(layers)} layer row(s) found. Verify against the source document."
        ),
        "warnings": warnings,
    }


def to_lab_data_format(parsed: Dict[str, Any]) -> Dict[str, Any]:
    """Converts parse_borehole_log_workbook()'s output into the same
    {"boreholes": {...}, "warnings": [...]} shape parse_uploaded_workbook()
    and the universal parser both return, so lab_data.py's
    parse_uploaded_workbook_auto() can treat all three paths identically."""
    meta = parsed["metadata"]
    borehole_id = meta.get("borehole_id") or parsed["sheet_name"]
    layers_out = []
    for l in parsed["layers"]:
        # Only fields that are real SoilLayer database columns (CANONICAL_FIELDS[k]["db"])
        # may reach db.add(SoilLayer(**layer_data)) -- anything else (liquid_limit,
        # plastic_limit, elastic_modulus, cbr, etc.) is a real soil property this
        # parser CAN recognize but the database schema has no column for, and passing
        # it through crashes the SQLAlchemy constructor outright.
        layer_out = {
            k: v for k, v in l.items()
            if k != "row" and CANONICAL_FIELDS.get(k, {}).get("db", False)
        }
        layers_out.append(layer_out)

    borehole = {
        "project_name": meta.get("project_name"),
        "project_number": meta.get("project_number"),
        "water_table_depth_m": _to_float(meta.get("water_table_depth_m")),
        "easting": _to_float(meta.get("easting")),
        "northing": _to_float(meta.get("northing")),
        "rl_m": _to_float(meta.get("rl_m")),
        "date_of_boring": meta.get("date_of_boring"),
        "layers": layers_out,
    }
    warnings = list(parsed["warnings"])
    warnings.append(parsed["confidence_note"])
    if not meta.get("water_table_depth_m"):
        warnings.append(f"Borehole {borehole_id}: water table depth not found in the sheet -- add it manually before running calculators.")

    return {"boreholes": {borehole_id: borehole}, "warnings": warnings}
