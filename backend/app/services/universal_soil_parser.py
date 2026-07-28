"""
universal_soil_parser.py
-------------------------
RaahiGeo AI - Universal Borehole/Soil-Investigation Excel Parser

Parses FLAT, row-per-sample soil investigation spreadsheets from ANY
consultant/lab -- regardless of exact column names, column order, units,
or a few title/logo rows above the real header -- into RaahiGeo's internal
schema (matching backend/app/models.py's BoreholeProfile + SoilLayer).

SCOPE / HONESTY NOTE:
This handles the common case: one header row (possibly not row 1, with
titles/logos/blank rows above it) followed by one row per soil layer/sample.
It does NOT handle multi-block "report style" workbooks with several
different data blocks scattered across 200+ columns at fixed offsets (that
is what bh_log_parser.py is for, for RaahiGeo's own office template). If a
sheet's real header row can't be found with reasonable confidence, this
parser says so explicitly rather than guessing.

WHAT THIS DELIBERATELY DOES NOT DO:
It never estimates/interpolates missing ENGINEERING VALUES (cohesion, phi,
Cc, density, etc.) from nearby layers -- soil properties can change sharply
between depths, and a wrong guess here could feed a wrong foundation design.
The only "recovery" performed is fill-down of STRUCTURAL fields that are
genuinely repeated-but-left-blank in the source sheet (Borehole ID, Project
Name, Water Table Depth, Easting/Northing, etc.) -- never a lab/field
measurement.

Usage:
    from universal_soil_parser import parse_workbook

    result = parse_workbook("/path/to/any_lab_report.xlsx")
    # result = {
    #   "sheet": "Sheet1",
    #   "header_row": 7,
    #   "column_mapping": {col_idx: {"field": ..., "header_text": ...,
    #                                  "confidence": 0-100,
    #                                  "unit_detected": ..., "unit_confidence": ...}},
    #   "unmapped_columns": [{"col_idx":..., "header_text":...}],
    #   "boreholes": {bh_id: {..fields.., "layers": [...]}},
    #   "extra_fields_not_in_db": {field_key: [values...]},  # detected but no DB column yet
    #   "warnings": [...],
    # }

Manual mapping memory (for columns this parser can't confidently guess):
    from universal_soil_parser import get_unmapped_columns, save_company_mapping

    unmapped = get_unmapped_columns(result)
    # ... show `unmapped` to the user, let them pick a field for each ...
    save_company_mapping("ACME Geotech Pvt Ltd", {"Unnamed: 14": "cohesion_t_m2"})
    # Next time a file is parsed with company_key="ACME Geotech Pvt Ltd",
    # that exact header text maps instantly at 100% confidence.
"""

from __future__ import annotations
import io
import re
import json
import difflib
import pathlib
from typing import Any, Dict, List, Optional, Union

import openpyxl

# ---------------------------------------------------------------------------
# Canonical field dictionary
# ---------------------------------------------------------------------------
# scope: "borehole" = one value per borehole (repeated down the sheet)
#        "layer"     = one value per soil layer / sample row
# db: True if backend/app/models.py already has a column for this (Aug 2026
#     schema). False = detected/extracted but not yet persisted to the DB --
#     surfaced under "extra_fields_not_in_db" instead of silently dropped.

CANONICAL_FIELDS: Dict[str, Dict[str, Any]] = {
    # ---- Borehole-level ----
    "borehole_id": {"scope": "borehole", "db": True, "synonyms": [
        "borehole id", "bh no", "bh id", "borehole no", "hole no", "bh number",
        "borehole number", "bh", "bore hole no", "borehole"]},
    "project_name": {"scope": "borehole", "db": True, "synonyms": [
        "project name", "project", "name of project", "name of work"]},
    "project_number": {"scope": "borehole", "db": True, "synonyms": [
        "project number", "job no", "job number", "project no", "work order no"]},
    "client_name": {"scope": "borehole", "db": False, "synonyms": [
        "client", "client name", "employer"]},
    "easting": {"scope": "borehole", "db": True, "synonyms": [
        "easting", "easting e", "e coordinate", "utm easting", "x coordinate", "x"]},
    "northing": {"scope": "borehole", "db": True, "synonyms": [
        "northing", "northing n", "n coordinate", "utm northing", "y coordinate", "y"]},
    "latitude": {"scope": "borehole", "db": False, "synonyms": ["latitude", "lat"]},
    "longitude": {"scope": "borehole", "db": False, "synonyms": ["longitude", "long", "lon"]},
    "rl_m": {"scope": "borehole", "db": True, "synonyms": [
        "rl", "reduced level", "ground level", "gl", "existing ground level",
        "surface level", "elevation", "collar elevation"]},
    "water_table_depth_m": {"scope": "borehole", "db": True, "synonyms": [
        "water table", "ground water level", "groundwater level",
        "water table depth", "gwl", "depth of water table", "standing water level"]},
    "date_of_boring": {"scope": "borehole", "db": True, "synonyms": [
        "date of boring", "boring date", "date", "date commenced", "commenced on"]},

    # ---- Layer-level: depth/identification ----
    "from_m": {"scope": "layer", "db": True, "synonyms": [
        "from", "depth from", "top depth", "from m", "layer from", "top"]},
    "to_m": {"scope": "layer", "db": True, "synonyms": [
        "to", "depth to", "bottom depth", "to m", "layer to", "bottom"]},
    "description": {"scope": "layer", "db": True, "synonyms": [
        "description", "soil description", "strata description",
        "description of strata", "soil type description"]},
    "classification": {"scope": "layer", "db": True, "synonyms": [
        "classification", "uscs", "is classification", "soil classification",
        "group symbol", "symbol", "uscs symbol"]},
    "sample_id": {"scope": "layer", "db": True, "synonyms": [
        "sample id", "sample no", "sample ref", "ref no", "sample reference"]},
    "sample_type": {"scope": "layer", "db": True, "synonyms": [
        "sample type", "type of sample"]},

    # ---- Layer-level: field/lab test values ----
    "n_value": {"scope": "layer", "db": True, "synonyms": [
        "spt", "spt n", "n value", "n60", "corrected spt", "corrected n",
        "field n", "blow count", "blows", "n"]},
    "bulk_density_t_m3": {"scope": "layer", "db": True, "unit_target": "t/m3", "synonyms": [
        "bulk density", "wet density", "moist density", "unit weight",
        "gamma", "bulk unit weight", "in situ density", "field density"]},
    "dry_density": {"scope": "layer", "db": False, "unit_target": "t/m3", "synonyms": [
        "dry density", "dry unit weight"]},
    "saturated_density": {"scope": "layer", "db": False, "unit_target": "t/m3", "synonyms": [
        "saturated density", "saturated unit weight"]},
    "specific_gravity": {"scope": "layer", "db": True, "synonyms": [
        "specific gravity", "sp gr", "gs", "g"]},
    "moisture_content_pct": {"scope": "layer", "db": True, "synonyms": [
        "moisture content", "natural moisture", "water content", "mc",
        "natural moisture content"]},
    "liquid_limit": {"scope": "layer", "db": False, "synonyms": ["liquid limit", "ll"]},
    "plastic_limit": {"scope": "layer", "db": False, "synonyms": ["plastic limit", "pl"]},
    "plasticity_index": {"scope": "layer", "db": False, "synonyms": [
        "plasticity index", "pi", "ip"]},
    "shrinkage_limit": {"scope": "layer", "db": False, "synonyms": ["shrinkage limit", "sl"]},
    "cohesion_t_m2": {"scope": "layer", "db": True, "unit_target": "t/m2", "synonyms": [
        "cohesion", "undrained shear strength", "cu", "c value", "cohesion intercept", "c"]},
    "friction_angle_deg": {"scope": "layer", "db": True, "synonyms": [
        "angle of internal friction", "phi", "friction angle",
        "angle of shearing resistance"]},
    "compression_index_cc": {"scope": "layer", "db": True, "synonyms": [
        "compression index", "cc"]},
    "recompression_index_cr": {"scope": "layer", "db": False, "synonyms": [
        "recompression index", "cr", "swelling index", "cs"]},
    "initial_void_ratio_e0": {"scope": "layer", "db": True, "synonyms": [
        "void ratio", "e0", "e", "initial void ratio", "initial e", "natural void ratio"]},
    "porosity": {"scope": "layer", "db": False, "synonyms": ["porosity"]},
    "relative_density": {"scope": "layer", "db": False, "synonyms": [
        "relative density", "dr"]},
    "degree_of_saturation": {"scope": "layer", "db": False, "synonyms": [
        "degree of saturation", "sr", "saturation"]},
    "coeff_consolidation_cv": {"scope": "layer", "db": False, "synonyms": [
        "coefficient of consolidation", "cv"]},
    "coeff_permeability_k": {"scope": "layer", "db": False, "synonyms": [
        "coefficient of permeability", "permeability", "k value", "hydraulic conductivity"]},
    "ocr": {"scope": "layer", "db": False, "synonyms": ["ocr", "overconsolidation ratio"]},
    "preconsolidation_pressure": {"scope": "layer", "db": False, "synonyms": [
        "preconsolidation pressure", "pc", "sigma c"]},
    "elastic_modulus": {"scope": "layer", "db": False, "synonyms": [
        "elastic modulus", "youngs modulus", "es", "modulus of elasticity"]},
    "poisson_ratio": {"scope": "layer", "db": False, "synonyms": [
        "poisson ratio", "poissons ratio", "nu"]},
    "cbr": {"scope": "layer", "db": False, "synonyms": ["cbr", "california bearing ratio"]},
    "ucs_kg_cm2": {"scope": "layer", "db": True, "unit_target": "kg/cm2", "synonyms": [
        "ucs", "unconfined compressive strength", "uniaxial compressive strength"]},
    "point_load_index": {"scope": "layer", "db": False, "synonyms": [
        "point load index", "is50", "pli"]},
    "rqd_pct": {"scope": "layer", "db": True, "synonyms": [
        "rqd", "rock quality designation"]},
    "core_recovery_pct": {"scope": "layer", "db": True, "synonyms": [
        "core recovery", "tcr", "recovery"]},
    "weathering_grade": {"scope": "layer", "db": True, "synonyms": [
        "weathering grade", "weathering", "grade of weathering"]},
    "rock_type": {"scope": "layer", "db": True, "synonyms": ["rock type", "lithology"]},
}

# ---------------------------------------------------------------------------
# Unit conversion tables (recognized_unit -> multiply raw value by this to
# get the internal target unit). Only fields with "unit_target" are checked.
# ---------------------------------------------------------------------------
UNIT_CONVERSIONS: Dict[str, Dict[str, float]] = {
    "t/m2": {"kpa": 0.101972, "kn/m2": 0.101972, "kg/cm2": 10.0, "mpa": 101.972, "t/m2": 1.0},
    "t/m3": {"g/cc": 1.0, "gm/cc": 1.0, "gcc": 1.0, "kg/m3": 0.001, "kn/m3": 1 / 9.80665, "t/m3": 1.0},
    "kg/cm2": {"kpa": 0.0101972, "mpa": 10.1972, "kg/cm2": 1.0, "t/m2": 0.1},
}
# Ordered so longer/more specific unit tokens are tried before short ones (e.g. "kn/m3" before "kn").
_UNIT_TOKEN_PATTERN = re.compile(
    r"(kn/m3|kn/m2|kg/cm2|kg/m3|t/m3|t/m2|g/cc|gm/cc|mpa|kpa)", re.IGNORECASE
)


def _normalize(text: Any) -> str:
    """Lowercase, strip units/punctuation/parens, collapse whitespace."""
    if text is None:
        return ""
    s = str(text).lower()
    s = re.sub(r"\(.*?\)", " ", s)          # drop parenthetical units e.g. "(kPa)"
    s = re.sub(r"[°'\"²³%]", " ", s)
    s = re.sub(r"[^a-z0-9\s]", " ", s)       # drop remaining punctuation
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _detect_unit(raw_header: str, field_key: str) -> Optional[Dict[str, Any]]:
    """Look for a known unit token in the raw header text for a field that has a unit_target."""
    target = CANONICAL_FIELDS.get(field_key, {}).get("unit_target")
    if not target:
        return None
    text = str(raw_header).lower().replace("\u00b2", "2").replace("\u00b3", "3")
    m = _UNIT_TOKEN_PATTERN.search(text)
    if not m:
        return {"unit": target, "factor": 1.0, "unit_confidence": 40,
                "note": "No unit found in header -- assumed already in internal unit; verify."}
    token = m.group(1).lower()
    factor = UNIT_CONVERSIONS.get(target, {}).get(token)
    if factor is None:
        return {"unit": token, "factor": 1.0, "unit_confidence": 30,
                "note": f"Unit '{token}' found but no conversion rule -- value used as-is; verify."}
    return {"unit": token, "factor": factor, "unit_confidence": 95}


def _score(header_norm: str, synonyms: List[str]) -> float:
    """Best match confidence (0-100) between a normalized header and a field's synonym list."""
    if not header_norm:
        return 0.0
    best = 0.0
    for syn in synonyms:
        if header_norm == syn:
            return 100.0
        if len(syn) <= 2:
            # Short synonyms like "x", "e", "g" are too ambiguous for substring/fuzzy
            # matching (e.g. "x" would wrongly match inside "XYZ column") -- exact only.
            continue
        if re.search(rf"\b{re.escape(syn)}\b", header_norm):
            best = max(best, 88.0)
        ratio = difflib.SequenceMatcher(None, header_norm, syn).ratio() * 100
        best = max(best, ratio)
    return best


def match_header(raw_header: Any) -> Optional[Dict[str, Any]]:
    """Return the best-matching canonical field for one raw header cell, or None."""
    norm = _normalize(raw_header)
    if not norm:
        return None
    best_field, best_conf = None, 0.0
    for field_key, spec in CANONICAL_FIELDS.items():
        conf = _score(norm, spec["synonyms"])
        if conf > best_conf:
            best_field, best_conf = field_key, conf
    if best_field is None or best_conf < 55:
        return None
    result = {"field": best_field, "confidence": round(best_conf, 1)}
    unit_info = _detect_unit(raw_header, best_field)
    if unit_info:
        result.update(unit_info)
    return result


# ---------------------------------------------------------------------------
# Header row auto-detection
# ---------------------------------------------------------------------------

def _find_header_row(ws, max_scan_rows: int = 40, min_fields_matched: int = 4):
    """Scan the first N rows for the one that looks most like a real header row."""
    best_row, best_score, best_map = None, 0.0, {}
    max_col = min(ws.max_column, 120)
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        col_map: Dict[int, Dict[str, Any]] = {}
        seen_fields: Dict[str, int] = {}
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            m = match_header(val)
            if not m:
                continue
            fk = m["field"]
            if fk in seen_fields and col_map[seen_fields[fk]]["confidence"] >= m["confidence"]:
                continue  # a better column for this field already found on this row
            m["col_idx"] = c
            m["header_text"] = val
            col_map[c] = m
            seen_fields[fk] = c
        distinct_fields = len(seen_fields)
        row_score = sum(v["confidence"] for v in col_map.values())
        if distinct_fields >= min_fields_matched and row_score > best_score:
            best_row, best_score, best_map = r, row_score, col_map
    return best_row, best_map


# ---------------------------------------------------------------------------
# Company-specific manual-mapping memory
# ---------------------------------------------------------------------------

_MAPPING_STORE_PATH = pathlib.Path(__file__).parent / "company_mappings.json"


def _load_store() -> Dict[str, Dict[str, str]]:
    if _MAPPING_STORE_PATH.exists():
        try:
            return json.loads(_MAPPING_STORE_PATH.read_text())
        except Exception:
            return {}
    return {}


def save_company_mapping(company_key: str, header_to_field: Dict[str, str]) -> None:
    """Persist manual header->field choices so future files from the same
    company are recognized automatically (at 100% confidence) next time."""
    store = _load_store()
    existing = store.get(company_key, {})
    existing.update({str(h): f for h, f in header_to_field.items()})
    store[company_key] = existing
    _MAPPING_STORE_PATH.write_text(json.dumps(store, indent=2))


def _company_override(raw_header: Any, company_key: Optional[str]) -> Optional[Dict[str, Any]]:
    if not company_key:
        return None
    store = _load_store()
    mapping = store.get(company_key, {})
    field = mapping.get(str(raw_header))
    if field:
        return {"field": field, "confidence": 100.0, "header_text": raw_header,
                "note": "From saved company mapping"}
    return None


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_workbook(file_path_or_bytes: Union[str, bytes],
                    company_key: Optional[str] = None,
                    manual_overrides: Optional[Dict[int, str]] = None) -> Dict[str, Any]:
    """
    Parse ANY flat, row-per-sample soil-investigation Excel sheet.

    company_key: an identifier (e.g. the consultant/lab's name) used to look
        up previously-saved manual column mappings for files from that source.
    manual_overrides: {col_idx: field_key} to force specific columns for THIS
        parse call (e.g. right after the user manually mapped some unmapped
        columns in the UI) -- combine with save_company_mapping() to persist.
    """
    warnings: List[str] = []
    source = io.BytesIO(file_path_or_bytes) if isinstance(file_path_or_bytes, bytes) else file_path_or_bytes
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    if len(wb.sheetnames) > 1:
        warnings.append(f"Workbook has {len(wb.sheetnames)} sheets -- only the first "
                         f"('{ws.title}') was parsed. Tell RaahiGeo which sheet to use if this is wrong.")

    header_row, column_mapping = _find_header_row(ws)
    if header_row is None:
        return {
            "sheet": ws.title, "header_row": None, "column_mapping": {},
            "unmapped_columns": [], "boreholes": {}, "extra_fields_not_in_db": {},
            "warnings": ["Could not confidently locate a header row in the first 40 rows. "
                         "This file's layout may need RaahiGeo's dedicated office-template "
                         "parser instead, or manual column mapping."],
        }

    # Apply company-saved mappings and per-call manual overrides (highest priority)
    max_col = min(ws.max_column, 120)
    raw_headers: Dict[int, Any] = {c: ws.cell(row=header_row, column=c).value for c in range(1, max_col + 1)}

    for c, val in raw_headers.items():
        if val is None or str(val).strip() == "":
            continue
        override = _company_override(val, company_key)
        if override:
            override["col_idx"] = c
            column_mapping[c] = override

    if manual_overrides:
        for c, field_key in manual_overrides.items():
            column_mapping[c] = {"field": field_key, "confidence": 100.0,
                                  "col_idx": c, "header_text": raw_headers.get(c),
                                  "note": "Manually mapped for this upload"}

    # Resolve duplicate columns mapped to the same field: keep highest confidence
    by_field: Dict[str, int] = {}
    duplicate_cols: List[int] = []
    for c, m in sorted(column_mapping.items()):
        fk = m["field"]
        if fk in by_field:
            if m["confidence"] > column_mapping[by_field[fk]]["confidence"]:
                duplicate_cols.append(by_field[fk])
                by_field[fk] = c
            else:
                duplicate_cols.append(c)
        else:
            by_field[fk] = c
    for c in duplicate_cols:
        column_mapping.pop(c, None)
    if duplicate_cols:
        warnings.append(f"{len(duplicate_cols)} column(s) were dropped as duplicates of a "
                         f"higher-confidence match for the same field (columns: {duplicate_cols}).")

    unmapped_columns = []
    for c, val in raw_headers.items():
        if val is None or str(val).strip() == "":
            continue
        if c not in column_mapping:
            unmapped_columns.append({"col_idx": c, "header_text": val})

    # ---- Extract data rows ----
    blank_streak = 0
    last_borehole_values: Dict[str, Any] = {}
    boreholes: Dict[str, Any] = {}
    extra_fields_not_in_db: Dict[str, List[Any]] = {}
    default_bh_counter = 0

    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = {c: ws.cell(row=r, column=c).value for c in column_mapping}
        if all(v in (None, "") for v in row_vals.values()):
            blank_streak += 1
            if blank_streak > 5:
                break
            continue
        blank_streak = 0

        record: Dict[str, Any] = {}
        for c, m in column_mapping.items():
            raw_val = row_vals.get(c)
            field_key = m["field"]
            factor = m.get("factor", 1.0)
            if raw_val is None or raw_val == "":
                value = None
            elif isinstance(raw_val, (int, float)) and factor != 1.0:
                value = raw_val * factor
            else:
                value = raw_val
            record[field_key] = value

        # Fill-down for STRUCTURAL (borehole-scope) fields only -- never for
        # layer-scope engineering values. This mirrors sheets where the
        # borehole ID/project/etc. is only typed once and left blank below.
        for field_key, spec in CANONICAL_FIELDS.items():
            if spec["scope"] != "borehole":
                continue
            if record.get(field_key) in (None, ""):
                if field_key in last_borehole_values:
                    record[field_key] = last_borehole_values[field_key]
            else:
                last_borehole_values[field_key] = record[field_key]

        bh_id = record.get("borehole_id")
        if not bh_id:
            default_bh_counter = default_bh_counter or 1
            bh_id = f"BH-UNKNOWN-{default_bh_counter}"
            warnings_msg = "No Borehole ID column recognized/found -- all rows grouped under a placeholder ID."
            if warnings_msg not in warnings:
                warnings.append(warnings_msg)

        if bh_id not in boreholes:
            bh_record = {k: record.get(k) for k, spec in CANONICAL_FIELDS.items() if spec["scope"] == "borehole"}
            bh_record["layers"] = []
            boreholes[bh_id] = bh_record

        layer_record = {k: record.get(k) for k, spec in CANONICAL_FIELDS.items()
                         if spec["scope"] == "layer" and spec["db"]}
        if layer_record.get("from_m") is None or layer_record.get("to_m") is None:
            warnings.append(f"Row {r}: skipped -- no From/To depth recognized.")
            continue
        boreholes[bh_id]["layers"].append(layer_record)

        # Surface non-DB fields separately rather than silently dropping them
        for k, spec in CANONICAL_FIELDS.items():
            if spec["scope"] == "layer" and not spec["db"] and record.get(k) is not None:
                extra_fields_not_in_db.setdefault(k, []).append(
                    {"row": r, "borehole_id": bh_id, "value": record[k]})

    result = {
        "sheet": ws.title,
        "header_row": header_row,
        "column_mapping": {c: {k: v for k, v in m.items() if k != "col_idx"} for c, m in column_mapping.items()},
        "unmapped_columns": unmapped_columns,
        "boreholes": boreholes,
        "extra_fields_not_in_db": extra_fields_not_in_db,
        "warnings": warnings,
        "low_confidence_overall": False,
    }

    # Sanity check: a real flat sheet usually has few boreholes with many
    # layers each. Many boreholes with very few layers each (or a huge
    # unmapped-column count relative to mapped) usually means this is a
    # multi-block "report style" workbook this parser isn't designed for,
    # NOT that the file genuinely has that many tiny boreholes.
    if boreholes:
        avg_layers = sum(len(b["layers"]) for b in boreholes.values()) / len(boreholes)
        if len(boreholes) > 5 and avg_layers < 3:
            result["low_confidence_overall"] = True
            result["warnings"].append(
                f"LOW CONFIDENCE: detected {len(boreholes)} boreholes averaging only "
                f"{avg_layers:.1f} layer(s) each -- this usually means the sheet is a "
                "multi-block 'report style' workbook (several data blocks at fixed "
                "column offsets) rather than a flat row-per-sample table, which this "
                "universal parser is not designed for. Don't trust this result -- use "
                "RaahiGeo's dedicated office-template parser (bh_log_parser.py) for "
                "this file, or map it manually."
            )
    if unmapped_columns and column_mapping and len(unmapped_columns) > 2 * len(column_mapping):
        result["low_confidence_overall"] = True
        result["warnings"].append(
            f"LOW CONFIDENCE: {len(unmapped_columns)} unmapped columns vs only "
            f"{len(column_mapping)} mapped -- the detected header row may be wrong."
        )

    return result


def get_unmapped_columns(parsed_result: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Convenience accessor -- what to show the user for one-time manual mapping."""
    return parsed_result.get("unmapped_columns", [])


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python universal_soil_parser.py <path_to_xlsx> [company_key]")
        sys.exit(1)
    company = sys.argv[2] if len(sys.argv) > 2 else None
    out = parse_workbook(sys.argv[1], company_key=company)
    print(json.dumps(out, indent=2, default=str))
