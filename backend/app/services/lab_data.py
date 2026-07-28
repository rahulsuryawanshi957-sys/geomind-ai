"""
Lab data import: generates the standardized Excel template, and parses a
filled-in copy back into structured BoreholeProfile + SoilLayer rows.

Template columns are denormalized (Borehole ID / Project / Water Table
repeated on every row) deliberately -- it keeps the sheet flat and simple
to fill in Excel, with no merged cells or multiple tabs to fight with.
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.config import logger

COLUMNS = [
    ("Borehole ID", "borehole_id", str),
    ("Project Name", "project_name", str),
    ("Water Table Depth (m)", "water_table_depth_m", float),
    ("From (m)", "from_m", float),
    ("To (m)", "to_m", float),
    ("Description", "description", str),
    ("Classification (USCS)", "classification", str),
    ("SPT N (field)", "n_value", float),
    ("Fines Content (%)", "fines_content_pct", float),
    ("Bulk Density (t/m3)", "bulk_density_t_m3", float),
    ("Specific Gravity", "specific_gravity", float),
    ("Moisture Content (%)", "moisture_content_pct", float),
    ("Cohesion C (t/m2)", "cohesion_t_m2", float),
    ("Friction Angle phi (deg)", "friction_angle_deg", float),
    ("Compression Index Cc", "compression_index_cc", float),
    ("Initial Void Ratio e0", "initial_void_ratio_e0", float),
    ("Rock Type", "rock_type", str),
    ("Weathering Grade", "weathering_grade", str),
    ("Core Recovery (%)", "core_recovery_pct", float),
    ("RQD (%)", "rqd_pct", float),
    ("UCS (kg/cm2)", "ucs_kg_cm2", float),
    ("Easting", "easting", float),
    ("Northing", "northing", float),
    ("R.L (m)", "rl_m", float),
    ("Date of Boring", "date_of_boring", str),
    ("Project Number", "project_number", str),
    ("Sample ID", "sample_id", str),
    ("Sample Type", "sample_type", str),
]

EXAMPLE_ROWS = [
    ["BH-01", "Sample Project", 3.5, 0, 1.5, "Filled up", "", "", "", 1.8, "", "", "", "", "", "", "", "", "", "", ""],
    ["BH-01", "Sample Project", 3.5, 1.5, 4.5, "Stiff silty clay", "CI", 8, 90, 1.9, 2.68, 22, 2.5, 0, 0.17, 0.75, "", "", "", "", ""],
    ["BH-01", "Sample Project", 3.5, 4.5, 10, "Medium dense sand", "SM", 18, 15, 1.85, 2.65, 15, 0, 30, "", "", "", "", "", "", ""],
    ["BH-02", "Sample Project", "", 0, 1.5, "Highly weathered fine-grained basalt", "", "", "", "", "", "", "", "", "", "", "Fine-Grained Basalt", "Grade IV", 30, 7, 120],
    ["BH-02", "Sample Project", "", 1.5, 3.0, "Fresh fine-grained basalt", "", "", "", "", "", "", "", "", "", "", "Fine-Grained Basalt", "Grade I", 95, 95, 850],
]


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Soil Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

    for col_idx, (label, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(14, len(label) // 1.3)

    for row_idx, row_data in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)

    notes_ws = wb.create_sheet("Instructions")
    notes = [
        "How to fill this sheet:",
        "- One row per soil layer (depth interval) for each borehole.",
        "- Repeat Borehole ID, Project Name, and Water Table Depth on every row for that borehole -- this is intentional.",
        "- Leave a cell blank if that value doesn't apply to the layer (e.g. Cohesion for a sandy layer).",
        "- Classification should be the USCS group symbol (CI, CL, SM, GW, etc.) where known.",
        "- Add as many boreholes as you like -- just continue adding rows with a new Borehole ID.",
        "- Do not rename or reorder the header columns in row 1 of the 'Soil Data' sheet.",
    ]
    for i, line in enumerate(notes, start=1):
        notes_ws.cell(row=i, column=1, value=line)
    notes_ws.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_uploaded_workbook(file_bytes: bytes) -> dict:
    """
    Returns {"boreholes": {borehole_id: {project_name, water_table_depth_m, layers: [...]}}, "warnings": [...]}
    Skips/warns on bad rows instead of failing the whole upload.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Soil Data" not in wb.sheetnames:
        raise ValueError("Expected a sheet named 'Soil Data' -- did you use the downloaded template?")
    ws = wb["Soil Data"]

    header_row = [cell.value for cell in ws[1]]
    expected_headers = [c[0] for c in COLUMNS]

    # Match columns BY NAME, not by strict position/full-set equality. An
    # older or externally-built sheet (e.g. Raahi's own BH_Log_Converter tool,
    # built before "Fines Content (%)" was added for Liquefaction Analysis)
    # would otherwise hard-fail on EVERY row just because one non-essential
    # column is missing or the columns are in a different order. Only the
    # columns actually needed to build a soil layer at all are required;
    # anything else missing is filled blank with a warning, same as an
    # individual blank cell already was.
    REQUIRED_HEADERS = {"Borehole ID", "Project Name", "Water Table Depth (m)", "From (m)", "To (m)"}
    header_to_col = {h: i for i, h in enumerate(header_row) if h}
    missing_headers = [h for h in expected_headers if h not in header_to_col]
    missing_required = [h for h in missing_headers if h in REQUIRED_HEADERS]
    if missing_required:
        raise ValueError(
            "Missing required column(s): " + ", ".join(missing_required) + ". "
            "Please use the downloaded template, or keep these exact column names if using your own sheet."
        )

    boreholes: dict = {}
    warnings = []
    if missing_headers:
        warnings.append(
            "This sheet is missing column(s) not present in the current template ("
            + ", ".join(missing_headers) + ") -- treated as blank for every row. "
            "Re-download the template if you want these filled in."
        )

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        row_dict = {}
        for label, key, cast in COLUMNS:
            col_idx = header_to_col.get(label)
            raw_value = row[col_idx] if col_idx is not None and col_idx < len(row) else None
            if raw_value is None or raw_value == "":
                row_dict[key] = None
                continue
            try:
                row_dict[key] = cast(raw_value)
            except (ValueError, TypeError):
                warnings.append(f"Row {row_num}: could not read '{label}' value {raw_value!r} -- left blank.")
                row_dict[key] = None

        bh_id = row_dict.get("borehole_id")
        if not bh_id:
            warnings.append(f"Row {row_num}: skipped -- missing Borehole ID.")
            continue
        if row_dict.get("from_m") is None or row_dict.get("to_m") is None:
            warnings.append(f"Row {row_num} (borehole {bh_id}): skipped -- missing From/To depth.")
            continue

        if bh_id not in boreholes:
            boreholes[bh_id] = {
                "project_name": row_dict.get("project_name"),
                "water_table_depth_m": row_dict.get("water_table_depth_m"),
                "easting": row_dict.get("easting"),
                "northing": row_dict.get("northing"),
                "rl_m": row_dict.get("rl_m"),
                "date_of_boring": row_dict.get("date_of_boring"),
                "project_number": row_dict.get("project_number"),
                "layers": [],
            }
        boreholes[bh_id]["layers"].append({
            "from_m": row_dict["from_m"],
            "to_m": row_dict["to_m"],
            "description": row_dict.get("description"),
            "classification": row_dict.get("classification"),
            "n_value": row_dict.get("n_value"),
            "fines_content_pct": row_dict.get("fines_content_pct"),
            "bulk_density_t_m3": row_dict.get("bulk_density_t_m3"),
            "specific_gravity": row_dict.get("specific_gravity"),
            "moisture_content_pct": row_dict.get("moisture_content_pct"),
            "cohesion_t_m2": row_dict.get("cohesion_t_m2"),
            "friction_angle_deg": row_dict.get("friction_angle_deg"),
            "compression_index_cc": row_dict.get("compression_index_cc"),
            "initial_void_ratio_e0": row_dict.get("initial_void_ratio_e0"),
            "rock_type": row_dict.get("rock_type"),
            "weathering_grade": row_dict.get("weathering_grade"),
            "core_recovery_pct": row_dict.get("core_recovery_pct"),
            "rqd_pct": row_dict.get("rqd_pct"),
            "ucs_kg_cm2": row_dict.get("ucs_kg_cm2"),
            "sample_id": row_dict.get("sample_id"),
            "sample_type": row_dict.get("sample_type"),
        })

    if not boreholes:
        raise ValueError("No valid rows found in the sheet.")

    logger.info(f"[lab_data] Parsed {len(boreholes)} borehole(s) from uploaded sheet, {len(warnings)} warning(s).")
    return {"boreholes": boreholes, "warnings": warnings}


def parse_uploaded_workbook_auto(file_bytes: bytes) -> dict:
    """
    Three-tier fallback so RaahiGeo accepts soil-investigation sheets from
    ANY source, not just its own template:
      1. RaahiGeo's own flat "Soil Data" template (exact match, fastest).
      2. The office-style borehole-log .xlsm format (bh_log_parser.py) --
         RaahiGeo's own report-style template with fixed cell positions.
      3. universal_soil_parser.py -- a THIRD PARTY consultant/lab's sheet
         with unknown column names/order/units, matched via a synonym
         dictionary + fuzzy matching. Only tried if 1 and 2 both fail, since
         it's the least certain of the three.
    All three paths return the SAME dict shape:
    {"boreholes": {borehole_id: {project_name, water_table_depth_m, ..., layers:[...]}}, "warnings":[...]}
    """
    from openpyxl import load_workbook
    import io
    from app.services.bh_log_parser import parse_borehole_log_workbook, to_lab_data_format

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Soil Data" in wb.sheetnames:
        return parse_uploaded_workbook(file_bytes)

    logger.info("[lab_data] 'Soil Data' sheet not found -- trying office borehole-log format.")
    try:
        parsed = parse_borehole_log_workbook(file_bytes)
        return to_lab_data_format(parsed)
    except Exception as e:
        logger.info(f"[lab_data] Office borehole-log parser did not match ({e}) -- trying universal parser.")

    from app.services.universal_soil_parser import parse_workbook as parse_universal

    universal_result = parse_universal(file_bytes)
    if universal_result.get("low_confidence_overall") or not universal_result.get("boreholes"):
        raise ValueError(
            "Could not automatically recognize this file's layout. "
            + " ".join(universal_result.get("warnings", []))
            or "Could not automatically recognize this file's layout -- try RaahiGeo's "
               "downloadable template, or contact support to add manual column mapping."
        )

    logger.info(f"[lab_data] Universal parser matched {len(universal_result['boreholes'])} "
                f"borehole(s); {len(universal_result.get('unmapped_columns', []))} column(s) unmapped.")

    boreholes_out: dict = {}
    for bh_id, bh in universal_result["boreholes"].items():
        boreholes_out[bh_id] = {
            "project_name": bh.get("project_name"),
            "water_table_depth_m": bh.get("water_table_depth_m"),
            "easting": bh.get("easting"),
            "northing": bh.get("northing"),
            "rl_m": bh.get("rl_m"),
            "date_of_boring": bh.get("date_of_boring"),
            "project_number": bh.get("project_number"),
            "layers": bh["layers"],
        }
    return {"boreholes": boreholes_out, "warnings": universal_result.get("warnings", [])}
